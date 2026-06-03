import sys, io, re, uuid
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
import openpyxl

wb_loan = openpyxl.load_workbook(r'C:\Users\jj_no\Desktop\coop-app-plan\April 2026 loan.xlsx', data_only=True)
wb_div  = openpyxl.load_workbook(r'C:\Users\jj_no\Desktop\coop-app-plan\Dividend Share 2026.xlsx', data_only=True)

# 1. Dividend sheet -> member list + capital
ws_div = wb_div['Dividend 2026']
div_rows = list(ws_div.iter_rows(values_only=True))

OWNER_THRESHOLD = 200_000
raw_members = []
for row in div_rows[2:]:
    name = row[0]
    capital = row[1]
    if not name or not isinstance(capital, (int, float)):
        continue
    name = str(name).strip()
    if name in ('put up', 'RCC COOP CAPITAL 2025-2026'):
        continue
    raw_members.append({'raw_name': name, 'capital': float(capital)})

print(f"Dividend members found: {len(raw_members)}", file=sys.stderr)

# 2. SUM LOAN -> loan balances
ws_loan = wb_loan['SUM LOAN']
loan_balances = {}
for row in ws_loan.iter_rows(values_only=True):
    name  = row[1]
    balance = row[6]
    beg = row[2]
    if name and isinstance(balance, (int, float)) and balance > 0 and name != 'TOTAL':
        loan_balances[str(name).strip()] = {'beginning': float(beg or 0), 'balance': float(balance)}

print(f"Active loan accounts: {len(loan_balances)}", file=sys.stderr)

# 3. Name parsing
def clean_name(raw):
    s = str(raw)
    s = re.sub(r'\s*-\s*ahongs.*$', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\s*c/o\s+.*', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\s*\(.*?\)', '', s)
    s = re.sub(r'\s+no\.\s*\d+$', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\s*-\s*(add|retired).+$', '', s, flags=re.IGNORECASE)
    if ',' in s:
        parts = s.split(',', 1)
        s = parts[1].strip() + ' ' + parts[0].strip()
    return s.strip()

def parse_name(raw):
    cleaned = clean_name(raw)
    parts = cleaned.split()
    if len(parts) == 0:
        return 'Unknown', 'Member'
    elif len(parts) == 1:
        return parts[0].title(), parts[0].title()
    else:
        surname = parts[0].title()
        firstname = ' '.join(parts[1:]).title()
        return surname, firstname

def make_slug(s):
    return re.sub(r'[^a-z0-9]', '', s.lower())

def make_email(surname, firstname):
    fn = make_slug(firstname.split()[0])
    sn = make_slug(surname)
    return f"{fn}.{sn}@rcccoop.com"

def make_password(surname):
    return make_slug(surname) + '1234'

# 4. Build member records
members = []
seen_emails = {}

for m in raw_members:
    surname, firstname = parse_name(m['raw_name'])
    email = make_email(surname, firstname)
    password = make_password(surname)

    base_email = email.replace('@rcccoop.com', '')
    if email in seen_emails:
        count = seen_emails.get(base_email, 1) + 1
        seen_emails[base_email] = count
        email = f"{base_email}{count}@rcccoop.com"
    else:
        seen_emails[email] = 1

    role = 'owner' if m['capital'] >= OWNER_THRESHOLD else 'member'
    mid = str(uuid.uuid4())
    uid = str(uuid.uuid4())

    members.append({
        'member_id': mid,
        'user_id':   uid,
        'raw_name':  m['raw_name'],
        'surname':   surname,
        'firstname': firstname,
        'full_name': f"{firstname} {surname}",
        'email':     email,
        'password':  password,
        'capital':   m['capital'],
        'role':      role,
        'loan':      None,
    })

print(f"Members to seed: {len(members)}", file=sys.stderr)

# 5. Match loans to members
def normalize(s):
    return re.sub(r'\s+', ' ', re.sub(r'[^a-z ]', '', str(s).lower())).strip()

loan_lookup = {normalize(k): v for k, v in loan_balances.items()}

for m in members:
    key1 = normalize(f"{m['surname']} {m['firstname']}")
    key2 = normalize(m['raw_name'])
    m['loan'] = loan_lookup.get(key1) or loan_lookup.get(key2)

matched = sum(1 for m in members if m['loan'])
print(f"Members with matched loans: {matched}", file=sys.stderr)

# 6. Build SQL
lines = []
lines.append("-- ============================================================")
lines.append("-- RCC COOP - Seed Data (April 2026 Excel records)")
lines.append("-- Run AFTER 001_initial_schema.sql in Supabase SQL Editor")
lines.append("--")
lines.append("-- IMPORTANT: Before running, disable email confirmation in")
lines.append("-- Supabase Dashboard > Authentication > Settings:")
lines.append("--   Toggle OFF: Enable email confirmations")
lines.append("-- ============================================================")
lines.append("")
lines.append("CREATE EXTENSION IF NOT EXISTS pgcrypto;")
lines.append("")

# auth.users
lines.append("-- Auth Users")
for m in members:
    email = m['email'].replace("'", "''")
    pw    = m['password'].replace("'", "''")
    uid   = m['user_id']
    lines.append(f"INSERT INTO auth.users (instance_id, id, aud, role, email, encrypted_password, email_confirmed_at, created_at, updated_at, raw_app_meta_data, raw_user_meta_data, is_super_admin, confirmation_token, email_change, email_change_token_new, recovery_token)")
    lines.append(f"VALUES ('00000000-0000-0000-0000-000000000000', '{uid}', 'authenticated', 'authenticated', '{email}', crypt('{pw}', gen_salt('bf')), NOW(), NOW(), NOW(), '{{\"provider\":\"email\",\"providers\":[\"email\"]}}', '{{}}', false, '', '', '', '');")

lines.append("")

# auth.identities
lines.append("-- Auth Identities")
for m in members:
    email = m['email'].replace("'", "''")
    uid   = m['user_id']
    iid   = str(uuid.uuid4())
    lines.append(f"INSERT INTO auth.identities (id, provider_id, user_id, identity_data, provider, created_at, updated_at)")
    lines.append(f"VALUES ('{iid}', '{uid}', '{uid}', '{{\"sub\":\"{uid}\",\"email\":\"{email}\",\"email_verified\":false,\"phone_verified\":false}}', 'email', NOW(), NOW());")

lines.append("")

# members table
lines.append("-- Members")
for m in members:
    name = m['full_name'].replace("'", "''")
    lines.append(f"INSERT INTO members (id, full_name, employment_type, membership_tier, date_joined, status, contribution_per_cutoff)")
    lines.append(f"VALUES ('{m['member_id']}', '{name}', 'regular', 'regular_member', '2020-01-01', 'active', 0);")

lines.append("")

# profiles
lines.append("-- Profiles")
for m in members:
    name = m['full_name'].replace("'", "''")
    lines.append(f"INSERT INTO profiles (id, member_id, role, display_name, first_login)")
    lines.append(f"VALUES ('{m['user_id']}', '{m['member_id']}', '{m['role']}', '{name}', true);")

lines.append("")

# capital contributions
lines.append("-- Share Capital")
for m in members:
    if m['capital'] and m['capital'] > 0:
        lines.append(f"INSERT INTO contributions (member_id, amount, cutoff_period, cutoff_date, payment_date, notes)")
        lines.append(f"VALUES ('{m['member_id']}', {m['capital']:.2f}, '2026-03-2ND', '2026-03-15', '2026-03-15', 'Capital seeded from FY2025-2026 dividend records');")

lines.append("")

# loans
lines.append("-- Outstanding Loans (balances as of April 2026)")
for m in members:
    if m['loan'] and m['loan']['balance'] > 0:
        bal = m['loan']['balance']
        lines.append(f"INSERT INTO loans (member_id, loan_number, loan_type, principal_amount, interest_rate, date_applied, date_released, status, notes)")
        lines.append(f"VALUES ('{m['member_id']}', 1, 'regular', {bal:.2f}, 2.00, '2020-01-01', '2020-01-01', 'released', 'Balance as of April 2026 seeded from Excel');")

lines.append("")
lines.append("-- ============================================================")
lines.append("-- CREDENTIAL REFERENCE — distribute to members, then delete this section")
lines.append("-- Name                              | Temp Email                                      | Temp Password")
for m in members:
    lines.append(f"-- {m['full_name']:<35} | {m['email']:<48} | {m['password']}")
lines.append("-- ============================================================")

output = '\n'.join(lines)
out_path = r'C:\Users\jj_no\Desktop\rcc-coop-app\supabase\migrations\002_seed_data.sql'
with open(out_path, 'w', encoding='utf-8') as f:
    f.write(output)

print(f"\nWritten to 002_seed_data.sql", file=sys.stderr)
print(f"Total members: {len(members)}", file=sys.stderr)
print(f"With loans:    {sum(1 for m in members if m['loan'])}", file=sys.stderr)
print(f"With capital:  {sum(1 for m in members if m['capital'] > 0)}", file=sys.stderr)
print(f"Owners:        {sum(1 for m in members if m['role'] == 'owner')}", file=sys.stderr)
